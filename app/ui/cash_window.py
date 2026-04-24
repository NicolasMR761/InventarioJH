from __future__ import annotations

from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QLineEdit,
    QComboBox,
    QDateEdit,
    QTextEdit,
    QDialog,
    QFileDialog,
    QFrame,
    QScrollArea,
    QHeaderView,
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor, QBrush, QWheelEvent

from app.db.cash_repo import (
    listar_movimientos,
    obtener_saldo,
    cerrar_dia,
    esta_cerrado,
    resumen_del_dia,
    resumen_rango,
    contar_movimientos,
)
from app.ui.cash_form import CashForm
from app.utils.formatters import fmt_fecha, fmt_cop as _fmt_cop


class _FixedTable(QTableWidget):
    """Tabla con scroll interno que no propaga la rueda al padre."""

    def wheelEvent(self, event: QWheelEvent):
        super().wheelEvent(event)
        event.accept()


class CashWindow(QWidget):
    def __init__(self):
        super().__init__()
        try:
            from app.main import get_icon

            if get_icon():
                self.setWindowIcon(get_icon())
        except Exception:
            pass
        self.setWindowTitle("Caja")
        self.resize(1050, 700)
        self.setStyleSheet(self._styles())

        self.page_size = 50
        self.offset = 0
        self.total_count = 0
        self._movs = []

        # ── Layout raíz con QScrollArea ──────────────
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        page_scroll = QScrollArea()
        page_scroll.setWidgetResizable(True)
        page_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        page_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        page_scroll.setFrameShape(QFrame.NoFrame)
        outer.addWidget(page_scroll)

        content = QWidget()
        page_scroll.setWidget(content)

        layout = QVBoxLayout(content)
        layout.setContentsMargins(18, 18, 18, 16)
        layout.setSpacing(10)

        # ── HEADER ────────────────────────────────────
        lbl_title = QLabel("🏦 Caja")
        lbl_title.setObjectName("pageTitle")
        lbl_sub = QLabel("Control de ingresos · Egresos · Cierre diario")
        lbl_sub.setObjectName("pageSub")
        layout.addWidget(lbl_title)
        layout.addWidget(lbl_sub)

        sep0 = QFrame()
        sep0.setFrameShape(QFrame.HLine)
        sep0.setObjectName("separator")
        layout.addWidget(sep0)

        # ── TARJETAS RESUMEN ──────────────────────────
        cards_row = QHBoxLayout()
        cards_row.setSpacing(12)
        self._card_saldo = self._make_card("SALDO PERÍODO", "$0,00")
        self._card_ing = self._make_card("INGRESOS", "$0,00", color="#4ade80")
        self._card_egr = self._make_card("EGRESOS", "$0,00", color="#f87171")
        self._card_estado = self._make_card("ESTADO", "—", color="#4ade80")
        cards_row.addWidget(self._card_saldo)
        cards_row.addWidget(self._card_ing)
        cards_row.addWidget(self._card_egr)
        cards_row.addWidget(self._card_estado)
        layout.addLayout(cards_row)

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.HLine)
        sep1.setObjectName("separator")
        layout.addWidget(sep1)

        # ── FILTROS ───────────────────────────────────
        filters = QHBoxLayout()
        filters.setSpacing(8)

        lbl_d = QLabel("Desde:")
        lbl_d.setObjectName("fieldLabel")
        filters.addWidget(lbl_d)
        self.dt_desde = QDateEdit()
        self.dt_desde.setObjectName("spinBox")
        self.dt_desde.setCalendarPopup(True)
        self.dt_desde.setDate(QDate.currentDate().addDays(-30))
        self.dt_desde.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.dt_desde)

        lbl_h = QLabel("Hasta:")
        lbl_h.setObjectName("fieldLabel")
        filters.addWidget(lbl_h)
        self.dt_hasta = QDateEdit()
        self.dt_hasta.setObjectName("spinBox")
        self.dt_hasta.setCalendarPopup(True)
        self.dt_hasta.setDate(QDate.currentDate())
        self.dt_hasta.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.dt_hasta)

        lbl_t = QLabel("Tipo:")
        lbl_t.setObjectName("fieldLabel")
        filters.addWidget(lbl_t)
        self.cbo_tipo = QComboBox()
        self.cbo_tipo.setObjectName("combo")
        self.cbo_tipo.addItems(["TODOS", "INGRESO", "EGRESO"])
        self.cbo_tipo.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.cbo_tipo)

        self.txt_buscar = QLineEdit()
        self.txt_buscar.setObjectName("inputField")
        self.txt_buscar.setPlaceholderText("Buscar concepto, referencia, observación…")
        filters.addWidget(self.txt_buscar, 2)

        btn_hoy = QPushButton("Hoy")
        btn_hoy.setObjectName("btnSecondary")
        btn_hoy.setCursor(Qt.ArrowCursor)
        btn_hoy.clicked.connect(self._filtro_hoy)
        filters.addWidget(btn_hoy)

        btn_aplicar = QPushButton("Aplicar")
        btn_aplicar.setObjectName("btnPrimary")
        btn_aplicar.setCursor(Qt.ArrowCursor)
        btn_aplicar.clicked.connect(lambda: self.cargar(reset_offset=True))
        filters.addWidget(btn_aplicar)

        btn_refrescar = QPushButton("↺")
        btn_refrescar.setObjectName("btnSecondary")
        btn_refrescar.setFixedWidth(38)
        btn_refrescar.setCursor(Qt.ArrowCursor)
        btn_refrescar.clicked.connect(lambda: self.cargar(reset_offset=False))
        filters.addWidget(btn_refrescar)

        layout.addLayout(filters)

        # ── ACCIONES ──────────────────────────────────
        actions = QHBoxLayout()
        actions.setSpacing(8)

        self.btn_ingreso = QPushButton("＋  Nuevo Ingreso")
        self.btn_ingreso.setObjectName("btnAccent")
        self.btn_ingreso.setCursor(Qt.ArrowCursor)
        self.btn_ingreso.clicked.connect(self.abrir_ingreso)

        self.btn_egreso = QPushButton("－  Nuevo Egreso")
        self.btn_egreso.setObjectName("btnAccent")
        self.btn_egreso.setCursor(Qt.ArrowCursor)
        self.btn_egreso.clicked.connect(self.abrir_egreso)

        actions.addWidget(self.btn_ingreso)
        actions.addWidget(self.btn_egreso)
        actions.addStretch()

        self.btn_export = QPushButton("↓ PDF")
        self.btn_export.setObjectName("btnSecondary")
        self.btn_export.setCursor(Qt.ArrowCursor)
        self.btn_export.clicked.connect(self.exportar_pdf)

        self.btn_excel = QPushButton("↓ Excel")
        self.btn_excel.setObjectName("btnSecondary")
        self.btn_excel.setCursor(Qt.ArrowCursor)
        self.btn_excel.clicked.connect(self.exportar_excel)

        self.btn_cierre = QPushButton("🔒 Cierre del día")
        self.btn_cierre.setObjectName("btnCierre")
        self.btn_cierre.setCursor(Qt.ArrowCursor)
        self.btn_cierre.clicked.connect(self.cerrar_dia_ui)

        actions.addWidget(self.btn_export)
        actions.addWidget(self.btn_excel)
        actions.addWidget(self.btn_cierre)
        layout.addLayout(actions)

        # ── TABLA ─────────────────────────────────────
        self.table = _FixedTable(0, 7)
        self.table.setObjectName("innerTable")
        self.table.setHorizontalHeaderLabels(
            ["ID", "Fecha", "Tipo", "Concepto", "Monto", "Referencia", "Detalle"]
        )
        self.table.setSortingEnabled(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(True)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.viewport().setCursor(Qt.ArrowCursor)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Fixed)
        hh.setSectionResizeMode(1, QHeaderView.Fixed)
        hh.setSectionResizeMode(2, QHeaderView.Fixed)
        hh.setSectionResizeMode(3, QHeaderView.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.Fixed)
        hh.setSectionResizeMode(5, QHeaderView.Fixed)
        hh.setSectionResizeMode(6, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 145)
        self.table.setColumnWidth(2, 90)
        self.table.setColumnWidth(4, 130)
        self.table.setColumnWidth(5, 120)
        self.table.setColumnWidth(6, 110)
        self.table.setFixedHeight(480)
        layout.addWidget(self.table)

        # ── PAGINACIÓN ────────────────────────────────
        pager = QHBoxLayout()
        self.btn_prev = QPushButton("« Anterior")
        self.btn_prev.setObjectName("btnSecondary")
        self.btn_prev.setFixedWidth(100)
        self.btn_prev.setCursor(Qt.ArrowCursor)
        self.btn_prev.clicked.connect(self.pagina_anterior)

        self.lbl_pager = QLabel("—")
        self.lbl_pager.setObjectName("pagerLabel")
        self.lbl_pager.setAlignment(Qt.AlignCenter)

        self.btn_next = QPushButton("Siguiente »")
        self.btn_next.setObjectName("btnSecondary")
        self.btn_next.setFixedWidth(100)
        self.btn_next.setCursor(Qt.ArrowCursor)
        self.btn_next.clicked.connect(self.pagina_siguiente)

        pager.addWidget(self.btn_prev)
        pager.addWidget(self.lbl_pager, 1)
        pager.addWidget(self.btn_next)
        layout.addLayout(pager)

        self.cargar(reset_offset=True)

    # ── CARDS ─────────────────────────────────────────
    def _make_card(self, label: str, value: str, color: str = "#93c5fd") -> QFrame:
        card = QFrame()
        card.setObjectName("summaryCard")
        card.setMinimumWidth(180)
        vl = QVBoxLayout(card)
        vl.setContentsMargins(16, 12, 16, 12)
        vl.setSpacing(4)
        lbl = QLabel(label)
        lbl.setObjectName("cardLabel")
        val = QLabel(value)
        val.setStyleSheet(
            f"font-size: 20px; font-weight: 900; color: {color}; letter-spacing: -0.5px;"
        )
        vl.addWidget(lbl)
        vl.addWidget(val)
        card._value_label = val
        card._color = color
        return card

    def _set_card(self, card: QFrame, value: str):
        card._value_label.setText(value)

    # ── ESTILOS ───────────────────────────────────────
    def _styles(self) -> str:
        return """
        QWidget {
            background: #0b1120;
            color: #e2e8f0;
            font-family: "Segoe UI", Arial, sans-serif;
            font-size: 13px;
        }
        QScrollArea { border: none; background: #0b1120; }

        #pageTitle {
            font-size: 20px; font-weight: 800;
            color: #f1f5f9; letter-spacing: -0.3px;
        }
        #pageSub  { font-size: 12px; color: #475569; }
        #fieldLabel { color: #64748b; font-size: 12px; font-weight: 600; }
        #pagerLabel { font-size: 11px; color: #475569; }
        #separator  { border: none; border-top: 1px solid #1e293b; }

        #summaryCard {
            background: #0b1120;
            border: 1px solid #1e3a5f;
            border-radius: 12px;
        }
        #cardLabel {
            font-size: 10px; font-weight: 700;
            color: #94a3b8; letter-spacing: 2px;
            text-transform: uppercase;
        }

        #inputField {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 6px 10px;
            color: #e2e8f0; min-height: 28px;
        }
        #inputField:focus { border-color: #3b82f6; }

        #spinBox {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 4px 8px;
            color: #e2e8f0; min-height: 28px;
        }

        #combo {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 4px 8px;
            color: #e2e8f0; min-height: 28px;
        }
        QComboBox::drop-down { border: none; width: 18px; }
        QComboBox QAbstractItemView {
            background: #111c33; border: 1px solid #1e3a5f;
            color: #e2e8f0; selection-background-color: #1e3a5f;
        }

        #btnPrimary {
            background: #2563eb; border: none; border-radius: 8px;
            padding: 6px 16px; font-weight: 700; color: white; min-height: 30px;
        }
        #btnPrimary:hover { background: #1d4ed8; }

        #btnSecondary {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 6px 12px;
            font-weight: 600; color: #94a3b8; min-height: 30px;
        }
        #btnSecondary:hover { border-color: #3b82f6; color: #e2e8f0; }
        #btnSecondary:disabled { color: #1e293b; border-color: #111c33; }

        #btnSuccess {
            background: #16a34a; border: none; border-radius: 8px;
            padding: 6px 18px; font-weight: 700; color: white; min-height: 30px;
        }
        #btnSuccess:hover { background: #15803d; }

        #btnDanger {
            background: #111c33; border: 1px solid #7f1d1d;
            border-radius: 8px; padding: 6px 12px;
            font-weight: 600; color: #f87171; min-height: 30px;
        }
        #btnDanger:hover { background: #1a0a0a; border-color: #ef4444; }

        #btnAccent {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 6px 18px;
            font-weight: 700; color: #93c5fd; min-height: 30px;
        }
        #btnAccent:hover { background: #1e3a5f; color: #fff; }

        #btnCierre {
            background: #111c33; border: 1px solid #92400e;
            border-radius: 8px; padding: 6px 14px;
            font-weight: 700; color: #fcd34d; min-height: 30px;
        }
        #btnCierre:hover { background: #1a1005; border-color: #f59e0b; }

        #innerTable {
            background: #0b1120;
            alternate-background-color: #0f1a2e;
            border: 1px solid #1e293b;
            border-radius: 8px;
            gridline-color: #1e293b;
            selection-background-color: #1e3a5f;
            selection-color: #f1f5f9;
            outline: none;
        }
        #innerTable QHeaderView::section {
            background: #111c33; color: #475569;
            font-size: 10px; font-weight: 700;
            letter-spacing: 1px; text-transform: uppercase;
            padding: 6px 10px;
            border: none; border-bottom: 2px solid #1e293b;
        }
        #innerTable::item { padding: 5px 10px; border: none; }
        #innerTable::item:selected { background: #1e3a5f; }

        QScrollBar:vertical {
            background: #0b1120; width: 6px; border-radius: 3px;
        }
        QScrollBar::handle:vertical { background: #1e3a5f; border-radius: 3px; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """

    # ── HELPERS ──────────────────────────────────────
    def _filtro_hoy(self):
        hoy = QDate.currentDate()
        self.dt_desde.setDate(hoy)
        self.dt_hasta.setDate(hoy)
        self.cargar(reset_offset=True)

    def _get_filters(self):
        d1 = self.dt_desde.date().toPython()
        d2 = self.dt_hasta.date().toPython()
        tipo = self.cbo_tipo.currentText()
        if tipo == "TODOS":
            tipo = None
        q = (self.txt_buscar.text() or "").strip() or None
        return d1, d2, tipo, q

    def _mov_by_id(self, mov_id: int):
        for m in self._movs:
            if int(getattr(m, "id", 0)) == int(mov_id):
                return m
        return None

    def _cell(self, txt: str, align=Qt.AlignLeft | Qt.AlignVCenter) -> QTableWidgetItem:
        it = QTableWidgetItem(str(txt))
        it.setTextAlignment(int(align))
        return it

    def _update_pager_ui(self):
        shown = len(self._movs)
        start = 0 if self.total_count == 0 else self.offset + 1
        end = self.offset + shown
        self.lbl_pager.setText(f"Mostrando {start}–{end} de {self.total_count}")
        self.btn_prev.setEnabled(self.offset > 0)
        self.btn_next.setEnabled(self.offset + self.page_size < self.total_count)

    # ── CARGA ─────────────────────────────────────────
    def keyPressEvent(self, event):
        from PySide6.QtCore import Qt

        key = event.key()
        mod = event.modifiers()
        if key == Qt.Key_F5:
            self.cargar(reset_offset=True)
        elif key == Qt.Key_Escape:
            self.close()
        elif key == Qt.Key_I and mod == Qt.ControlModifier:
            self.abrir_ingreso()
        elif key == Qt.Key_E and mod == Qt.ControlModifier:
            self.abrir_egreso()
        else:
            super().keyPressEvent(event)

    def cargar(self, reset_offset: bool = False):
        try:
            if reset_offset:
                self.offset = 0

            d1, d2, tipo, q = self._get_filters()

            if d1 == d2 and esta_cerrado(d1):
                self._set_card(self._card_estado, f"🔒 CERRADO")
            else:
                self._set_card(self._card_estado, "✅ ABIERTO")

            data = resumen_del_dia(d1) if d1 == d2 else resumen_rango(d1, d2)
            ingresos = float(data["ingresos"] or 0.0)
            egresos = float(data["egresos"] or 0.0)
            balance = ingresos - egresos

            self._set_card(self._card_saldo, _fmt_cop(balance))
            self._set_card(self._card_ing, _fmt_cop(ingresos))
            self._set_card(self._card_egr, _fmt_cop(egresos))

            self.total_count = contar_movimientos(
                fecha_desde=d1, fecha_hasta=d2, tipo=tipo, q=q
            )
            self._movs = listar_movimientos(
                limit=self.page_size,
                offset=self.offset,
                fecha_desde=d1,
                fecha_hasta=d2,
                tipo=tipo,
                q=q,
            )

            was_sort = self.table.isSortingEnabled()
            self.table.setSortingEnabled(False)
            self.table.blockSignals(True)
            self.table.setRowCount(len(self._movs))

            for row, m in enumerate(self._movs):
                self.table.setRowHeight(row, 32)
                fecha_txt = fmt_fecha(m.fecha) if getattr(m, "fecha", None) else ""
                es_ingreso = (m.tipo or "").upper() == "INGRESO"

                self.table.setItem(
                    row, 0, self._cell(str(m.id), Qt.AlignCenter | Qt.AlignVCenter)
                )
                self.table.setItem(row, 1, self._cell(fecha_txt))

                it_tipo = QTableWidgetItem(m.tipo or "")
                it_tipo.setForeground(
                    QBrush(QColor("#4ade80" if es_ingreso else "#f87171"))
                )
                it_tipo.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                self.table.setItem(row, 2, it_tipo)

                self.table.setItem(row, 3, self._cell(m.concepto or ""))

                it_m = QTableWidgetItem(_fmt_cop(m.monto or 0.0))
                it_m.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                it_m.setForeground(
                    QBrush(QColor("#4ade80" if es_ingreso else "#f87171"))
                )
                self.table.setItem(row, 4, it_m)

                self.table.setItem(row, 5, self._cell(m.referencia or ""))

                mov_id = int(m.id)
                btn = QPushButton("Ver detalle")
                btn.setCursor(Qt.ArrowCursor)
                btn.setStyleSheet(
                    """
                    QPushButton {
                        background: #111c33; border: 1px solid #1e3a5f;
                        border-radius: 6px; padding: 3px 10px;
                        color: #93c5fd; font-size: 12px; font-weight: 600;
                    }
                    QPushButton:hover { background: #1e3a5f; color: #fff; }
                """
                )
                btn.clicked.connect(
                    lambda _=False, mid=mov_id: self.ver_detalle_por_id(mid)
                )
                self.table.setCellWidget(row, 6, btn)

            self.table.blockSignals(False)
            self.table.setSortingEnabled(was_sort)
            self._update_pager_ui()

        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    # ── PAGINACIÓN ────────────────────────────────────
    def pagina_anterior(self):
        self.offset = max(0, self.offset - self.page_size)
        self.cargar(reset_offset=False)

    def pagina_siguiente(self):
        if self.offset + self.page_size < self.total_count:
            self.offset += self.page_size
        self.cargar(reset_offset=False)

    # ── ACCIONES ─────────────────────────────────────
    def abrir_ingreso(self):
        form = CashForm(tipo="INGRESO", parent=self)
        if form.exec():
            self.cargar(reset_offset=False)

    def abrir_egreso(self):
        form = CashForm(tipo="EGRESO", parent=self)
        if form.exec():
            self.cargar(reset_offset=False)

    # ── DETALLE ───────────────────────────────────────
    def ver_detalle_por_id(self, mov_id: int) -> None:
        m = self._mov_by_id(mov_id)
        if not m:
            QMessageBox.information(self, "Detalle", "No se encontró el movimiento.")
            return

        fecha_txt = ""
        if getattr(m, "fecha", None):
            try:
                fecha_txt = m.fecha.strftime("%d/%m/%Y")
            except Exception:
                fecha_txt = str(m.fecha)

        es_ingreso = (m.tipo or "").upper() == "INGRESO"
        color_tipo = "#22c55e" if es_ingreso else "#ef4444"
        label_tipo = "INGRESO" if es_ingreso else "EGRESO"
        emoji_tipo = "↑" if es_ingreso else "↓"

        obs_raw = (m.observacion or "").strip()
        obs_lines = obs_raw.split("\n") if obs_raw else []
        obs_html = (
            "".join(
                f"<tr><td style='padding:5px 0;color:#cbd5e1;font-size:13px;"
                f"border-bottom:1px solid #1e293b;'>{ln}</td></tr>"
                for ln in obs_lines
            )
            if obs_lines
            else "<tr><td style='color:#475569;font-size:12px;'>Sin detalle</td></tr>"
        )
        ref_txt = (m.referencia or "—").strip()

        html = f"""
        <html><body style='margin:0;padding:0;background:#0a0f1e;
                           font-family:"Segoe UI",Arial,sans-serif;'>
        <div style='background:#0a0f1e;'>
            <div style='background:linear-gradient(135deg,{color_tipo}cc,{color_tipo}66);
                        padding:20px 28px 16px 28px;'>
                <div style='font-size:10px;color:rgba(255,255,255,0.65);letter-spacing:3px;
                            text-transform:uppercase;margin-bottom:6px;'>
                    INVENTARIO JH &nbsp;·&nbsp; Comprobante Interno</div>
                <div style='font-size:24px;font-weight:800;color:#ffffff;'>
                    {m.concepto or "Sin concepto"}</div>
                <div style='margin-top:10px;'>
                    <span style='background:rgba(0,0,0,0.3);color:#fff;padding:3px 12px;
                                 border-radius:20px;font-size:11px;font-weight:700;
                                 letter-spacing:1.5px;'>{emoji_tipo} {label_tipo}</span>
                    <span style='color:rgba(255,255,255,0.55);font-size:11px;margin-left:8px;'>
                        Ref: {ref_txt}</span>
                </div>
            </div>
            <div style='background:#0f172a;padding:20px 28px;border-bottom:1px solid #1e293b;'>
                <div style='font-size:10px;color:#475569;letter-spacing:3px;
                            margin-bottom:4px;text-transform:uppercase;'>Monto total</div>
                <div style='font-size:36px;font-weight:900;color:{color_tipo};letter-spacing:-1px;'>
                    {_fmt_cop(m.monto or 0.0)}</div>
            </div>
            <div style='padding:16px 28px;border-bottom:1px solid #1e293b;'>
                <table width='100%' cellspacing='0' cellpadding='0'>
                    <tr><td style='padding:6px 0;color:#475569;font-size:11px;
                                   text-transform:uppercase;letter-spacing:1px;width:38%;'>
                            N° Registro</td>
                        <td style='padding:6px 0;color:#e2e8f0;font-size:13px;font-weight:600;'>
                            #{m.id}</td></tr>
                    <tr><td style='padding:6px 0;color:#475569;font-size:11px;
                                   text-transform:uppercase;letter-spacing:1px;'>Fecha</td>
                        <td style='padding:6px 0;color:#e2e8f0;font-size:13px;font-weight:600;'>
                            {fecha_txt}</td></tr>
                    <tr><td style='padding:6px 0;color:#475569;font-size:11px;
                                   text-transform:uppercase;letter-spacing:1px;'>Referencia</td>
                        <td style='padding:6px 0;color:#e2e8f0;font-size:13px;font-weight:600;'>
                            {ref_txt}</td></tr>
                </table>
            </div>
            <div style='padding:14px 28px 8px 28px;'>
                <div style='font-size:10px;color:#475569;letter-spacing:2px;
                            text-transform:uppercase;margin-bottom:10px;'>Detalle</div>
                <table width='100%' cellspacing='0' cellpadding='0'>{obs_html}</table>
            </div>
        </div></body></html>"""

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Comprobante #{m.id}")
        dlg.setFixedWidth(480)
        dlg.setStyleSheet("background:#0a0f1e;")
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(0, 0, 0, 12)
        lay.setSpacing(0)

        txt = QTextEdit()
        txt.setReadOnly(True)
        txt.setHtml(html)
        txt.setStyleSheet(
            """
            QTextEdit { background:#0a0f1e; border:none; color:#e2e8f0; }
            QScrollBar:vertical { background:#0a0f1e; width:6px; }
            QScrollBar::handle:vertical { background:#334155; border-radius:3px; }
        """
        )
        txt.setMinimumHeight(460)
        lay.addWidget(txt)

        btn_cerrar = QPushButton("Cerrar")
        btn_cerrar.setFixedWidth(120)
        btn_cerrar.setCursor(Qt.ArrowCursor)
        btn_cerrar.setStyleSheet(
            """
            QPushButton { background:#1e3a5f; color:#e2e8f0; border:1px solid #2563eb;
                          border-radius:6px; padding:7px 0; font-weight:700; }
            QPushButton:hover { background:#2563eb; }
        """
        )
        btn_cerrar.clicked.connect(dlg.accept)
        row_btn = QHBoxLayout()
        row_btn.addStretch()
        row_btn.addWidget(btn_cerrar)
        row_btn.addStretch()
        lay.addLayout(row_btn)
        dlg.exec()

    # ── EXPORTAR ─────────────────────────────────────
    def _movs_para_exportar(self):
        d1, d2, tipo, q = self._get_filters()
        return listar_movimientos(
            limit=200000, offset=0, fecha_desde=d1, fecha_hasta=d2, tipo=tipo, q=q
        )

    def exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(
                self,
                "Dependencia faltante",
                "Instala reportlab:\n\npip install reportlab",
            )
            return
        try:
            d1, d2, tipo, q = self._get_filters()
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar PDF", f"caja_{d1}_a_{d2}.pdf", "PDF (*.pdf)"
            )
            if not path:
                return
            movs = self._movs_para_exportar()
            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            y = h - 2 * cm
            c.setFont("Helvetica-Bold", 14)
            from app.utils.config_manager import cargar_config

            _cfg = cargar_config()
            _nombre_empresa = _cfg.get("empresa_nombre") or "Inventario JH"
            c.drawString(2 * cm, y, f"Reporte de Caja — {_nombre_empresa}")
            y -= 0.7 * cm
            c.setFont("Helvetica", 10)
            c.drawString(2 * cm, y, f"Rango: {d1} a {d2}")
            y -= 0.5 * cm
            c.drawString(2 * cm, y, f"Tipo: {tipo or 'TODOS'}   Buscar: {q or '-'}")
            y -= 0.8 * cm
            data = resumen_del_dia(d1) if d1 == d2 else resumen_rango(d1, d2)
            for m in movs:
                if y < 2 * cm:
                    c.showPage()
                    y = h - 2 * cm
                fecha_txt = fmt_fecha(m.fecha) if getattr(m, "fecha", None) else ""
                line = (m.concepto or "").strip()
                ref = (m.referencia or "").strip()
                if ref:
                    line += f" ({ref})"
                c.drawString(2 * cm, y, str(fecha_txt)[:16])
                c.drawString(6.2 * cm, y, (m.tipo or "")[:10])
                c.drawRightString(10.8 * cm, y, _fmt_cop(m.monto or 0.0))
                c.drawString(11.2 * cm, y, line[:60])
                y -= 0.38 * cm
            c.save()
            QMessageBox.information(self, "OK", "PDF exportado correctamente.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self,
                "Dependencia faltante",
                "Instala openpyxl:\n\npip install openpyxl",
            )
            return
        try:
            d1, d2, tipo, q = self._get_filters()
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar Excel", f"caja_{d1}_a_{d2}.xlsx", "Excel (*.xlsx)"
            )
            if not path:
                return
            movs = self._movs_para_exportar()
            wb = Workbook()
            ws = wb.active
            ws.title = "Caja"
            ws["A1"] = "Reporte de Caja"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = f"Rango: {d1} a {d2}"
            row_ptr = 4
            headers = [
                "ID",
                "Fecha",
                "Tipo",
                "Concepto",
                "Monto",
                "Referencia",
                "Observación",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_ptr, column=col, value=header)
                cell.font = Font(bold=True)
            row_ptr += 1
            for m in movs:
                ws.cell(row=row_ptr, column=1, value=int(m.id))
                ws.cell(row=row_ptr, column=2, value=fmt_fecha(m.fecha))
                ws.cell(row=row_ptr, column=3, value=m.tipo)
                ws.cell(row=row_ptr, column=4, value=m.concepto)
                ws.cell(row=row_ptr, column=5, value=float(m.monto or 0.0))
                ws.cell(row=row_ptr, column=5).number_format = "#,##0.00"
                ws.cell(row=row_ptr, column=6, value=m.referencia or "")
                ws.cell(row=row_ptr, column=7, value=m.observacion or "")
                row_ptr += 1
            for i, w in enumerate([10, 20, 12, 30, 14, 18, 30], 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            wb.save(path)
            QMessageBox.information(self, "OK", "Excel exportado correctamente.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def cerrar_dia_ui(self):
        try:
            d = self.dt_desde.date().toPython()
            d2 = self.dt_hasta.date().toPython()
            if d != d2:
                QMessageBox.warning(
                    self,
                    "Cierre",
                    "Para cierre diario, selecciona un solo día (Desde = Hasta).",
                )
                return
            if esta_cerrado(d):
                QMessageBox.information(self, "Cierre", f"El día {d} ya está cerrado.")
                return
            data = resumen_del_dia(d)
            msg = (
                f"Cierre del día: {d}\n\n"
                f"Saldo inicial : {_fmt_cop(data['saldo_inicial'])}\n"
                f"Ingresos      : {_fmt_cop(data['ingresos'])}\n"
                f"Egresos       : {_fmt_cop(data['egresos'])}\n"
                f"Saldo final   : {_fmt_cop(data['saldo_final'])}\n\n"
                f"¿Confirmas cerrar el día?"
            )
            if QMessageBox.question(self, "Confirmar cierre", msg) != QMessageBox.Yes:
                return
            cerrar_dia(d, cerrado_por=None)
            QMessageBox.information(self, "OK", f"Día {d} cerrado.")
            self.cargar(reset_offset=False)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
