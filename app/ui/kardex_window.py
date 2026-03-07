from __future__ import annotations

from datetime import datetime, time

from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QBrush, QColor, QFont, QWheelEvent
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QComboBox,
    QDateEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QFrame,
    QFileDialog,
    QMessageBox,
    QScrollArea,
)

from app.db.products_repo import listar_productos
from app.db.kardex_repo import obtener_kardex
from app.utils.formatters import fmt_fecha


def _fmt_qty(x: float) -> str:
    try:
        v = float(x or 0.0)
        s = f"{v:.3f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")
    except Exception:
        return "0"


def _fmt_cop(value: float) -> str:
    """Formato COP: sin decimales, sin signo negativo, separador de miles con punto."""
    try:
        n = int(round(abs(float(value or 0.0))))
        s = ""
        digits = str(n)
        for i, ch in enumerate(reversed(digits)):
            if i > 0 and i % 3 == 0:
                s = "." + s
            s = ch + s
        return "$" + s
    except Exception:
        return "$0"


def _item(
    text: str, align: Qt.AlignmentFlag | Qt.Alignment = Qt.AlignLeft | Qt.AlignVCenter
) -> QTableWidgetItem:
    it = QTableWidgetItem(text)
    it.setTextAlignment(int(align))
    return it


class _FixedTable(QTableWidget):
    def wheelEvent(self, event: QWheelEvent):
        super().wheelEvent(event)
        event.accept()


class KardexWindow(QWidget):
    def __init__(self):
        super().__init__()
        try:
            from app.main import get_icon

            if get_icon():
                self.setWindowIcon(get_icon())
        except Exception:
            pass
        self.setWindowTitle("Kardex por Producto")
        self.resize(1120, 720)
        self.setStyleSheet(self._styles())

        self._prod_unidad: dict[int, str] = {}

        # ── Layout raíz con QScrollArea ──────────────
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)

        page_scroll = QScrollArea()
        page_scroll.setWidgetResizable(True)
        page_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        page_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        page_scroll.setFrameShape(QFrame.NoFrame)
        outer.addWidget(page_scroll)

        content = QWidget()
        page_scroll.setWidget(content)

        layout = QVBoxLayout(content)
        layout.setContentsMargins(18, 18, 18, 16)
        layout.setSpacing(10)

        # ── HEADER ────────────────────────────────────
        lbl_title = QLabel("📦 Kardex por Producto")
        lbl_title.setObjectName("pageTitle")
        lbl_sub = QLabel("Movimientos de inventario · Entradas · Ventas · Anulaciones")
        lbl_sub.setObjectName("pageSub")
        layout.addWidget(lbl_title)
        layout.addWidget(lbl_sub)

        sep0 = QFrame()
        sep0.setFrameShape(QFrame.HLine)
        sep0.setObjectName("separator")
        layout.addWidget(sep0)

        # ── FILTROS ───────────────────────────────────
        filters = QHBoxLayout()
        filters.setSpacing(8)

        lbl_p = QLabel("Producto:")
        lbl_p.setObjectName("fieldLabel")
        filters.addWidget(lbl_p)
        self.cbo_producto = QComboBox()
        self.cbo_producto.setObjectName("combo")
        self.cbo_producto.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.cbo_producto, 2)

        lbl_d = QLabel("Desde:")
        lbl_d.setObjectName("fieldLabel")
        filters.addWidget(lbl_d)
        self.dt_desde = QDateEdit()
        self.dt_desde.setObjectName("spinBox")
        self.dt_desde.setCalendarPopup(True)
        self.dt_desde.setDisplayFormat("dd/MM/yyyy")
        self.dt_desde.setDate(QDate.currentDate().addDays(-7))
        self.dt_desde.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.dt_desde)

        lbl_h = QLabel("Hasta:")
        lbl_h.setObjectName("fieldLabel")
        filters.addWidget(lbl_h)
        self.dt_hasta = QDateEdit()
        self.dt_hasta.setObjectName("spinBox")
        self.dt_hasta.setCalendarPopup(True)
        self.dt_hasta.setDisplayFormat("dd/MM/yyyy")
        self.dt_hasta.setDate(QDate.currentDate())
        self.dt_hasta.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.dt_hasta)

        self.btn_cargar = QPushButton("Cargar")
        self.btn_cargar.setObjectName("btnPrimary")
        self.btn_cargar.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.btn_cargar)

        self.btn_pdf = QPushButton("↓ PDF")
        self.btn_pdf.setObjectName("btnSecondary")
        self.btn_pdf.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.btn_pdf)

        self.btn_excel = QPushButton("↓ Excel")
        self.btn_excel.setObjectName("btnSecondary")
        self.btn_excel.setCursor(Qt.ArrowCursor)
        filters.addWidget(self.btn_excel)

        layout.addLayout(filters)

        # ── INFO SALDO INICIAL ────────────────────────
        self.lbl_info = QLabel("Saldo inicial (antes del rango): 0")
        self.lbl_info.setObjectName("infoLabel")
        self.lbl_info.setAlignment(Qt.AlignLeft)
        self.lbl_info.setWordWrap(True)
        layout.addWidget(self.lbl_info)

        # ── TARJETAS RESUMEN ──────────────────────────
        cards_row = QHBoxLayout()
        cards_row.setSpacing(12)
        self._card_entradas = self._make_card("ENTRADAS", "0", color="#4ade80")
        self._card_ventas = self._make_card("VENTAS", "0", color="#f87171")
        self._card_anul = self._make_card("ANULACIONES", "0", color="#fbbf24")
        self._card_neto = self._make_card("NETO", "0")
        self._card_saldo = self._make_card("SALDO FINAL", "0", color="#93c5fd")
        cards_row.addWidget(self._card_entradas)
        cards_row.addWidget(self._card_ventas)
        cards_row.addWidget(self._card_anul)
        cards_row.addWidget(self._card_neto)
        cards_row.addWidget(self._card_saldo)
        layout.addLayout(cards_row)

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.HLine)
        sep1.setObjectName("separator")
        layout.addWidget(sep1)

        # ── TABLA ─────────────────────────────────────
        self.table = _FixedTable(0, 7)
        self.table.setObjectName("innerTable")
        self.table.setHorizontalHeaderLabels(
            ["Fecha", "Tipo", "Referencia", "Cantidad", "Precio", "Subtotal", "Saldo"]
        )
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(False)
        self.table.setSortingEnabled(False)
        self.table.viewport().setCursor(Qt.ArrowCursor)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)

        self.table.setFixedHeight(440)
        layout.addWidget(self.table)

        # ── SEÑALES ───────────────────────────────────
        self.btn_cargar.clicked.connect(self.cargar)
        self.btn_pdf.clicked.connect(self.exportar_pdf)
        self.btn_excel.clicked.connect(self.exportar_excel)

        self._cargar_productos()
        self.cargar()

    # ── CARDS ─────────────────────────────────────────
    def _make_card(self, label: str, value: str, color: str = "#e2e8f0") -> QFrame:
        card = QFrame()
        card.setObjectName("summaryCard")
        vl = QVBoxLayout(card)
        vl.setContentsMargins(14, 10, 14, 10)
        vl.setSpacing(3)
        lbl = QLabel(label)
        lbl.setObjectName("cardLabel")
        val = QLabel(value)
        val.setStyleSheet(
            f"font-size: 18px; font-weight: 900; color: {color}; letter-spacing: -0.5px;"
        )
        vl.addWidget(lbl)
        vl.addWidget(val)
        card._value_label = val
        return card

    def _set_card(self, card: QFrame, value: str):
        card._value_label.setText(value)

    # ── ESTILOS ───────────────────────────────────────
    def _styles(self) -> str:
        return """
        QWidget {
            background: #0b1120;
            color: #e2e8f0;
            font-family: "Segoe UI", Arial, sans-serif;
            font-size: 13px;
        }
        QScrollArea { border: none; background: #0b1120; }

        #pageTitle {
            font-size: 20px; font-weight: 800;
            color: #f1f5f9; letter-spacing: -0.3px;
        }
        #pageSub   { font-size: 12px; color: #475569; }
        #fieldLabel { color: #64748b; font-size: 12px; font-weight: 600; }
        #infoLabel  { font-size: 12px; color: #94a3b8; padding: 4px 0; }
        #separator  { border: none; border-top: 1px solid #1e293b; }

        #summaryCard {
            background: #0b1120;
            border: 1px solid #1e3a5f;
            border-radius: 12px;
        }
        #cardLabel {
            font-size: 10px; font-weight: 700;
            color: #94a3b8; letter-spacing: 2px;
            text-transform: uppercase;
        }

        #combo {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 4px 8px;
            color: #e2e8f0; min-height: 28px;
        }
        QComboBox::drop-down { border: none; width: 18px; }
        QComboBox QAbstractItemView {
            background: #111c33; border: 1px solid #1e3a5f;
            color: #e2e8f0; selection-background-color: #1e3a5f;
        }

        #spinBox {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 4px 8px;
            color: #e2e8f0; min-height: 28px;
        }

        #btnPrimary {
            background: #2563eb; border: none; border-radius: 8px;
            padding: 6px 16px; font-weight: 700; color: white; min-height: 30px;
        }
        #btnPrimary:hover { background: #1d4ed8; }

        #btnSecondary {
            background: #111c33; border: 1px solid #1e3a5f;
            border-radius: 8px; padding: 6px 12px;
            font-weight: 600; color: #94a3b8; min-height: 30px;
        }
        #btnSecondary:hover { border-color: #3b82f6; color: #e2e8f0; }

        #innerTable {
            background: #0b1120;
            alternate-background-color: #0f1a2e;
            border: 1px solid #1e293b;
            border-radius: 8px;
            gridline-color: #1e293b;
            selection-background-color: #1e3a5f;
            selection-color: #f1f5f9;
            outline: none;
        }
        #innerTable QHeaderView::section {
            background: #111c33; color: #475569;
            font-size: 10px; font-weight: 700;
            letter-spacing: 1px; text-transform: uppercase;
            padding: 6px 10px;
            border: none; border-bottom: 2px solid #1e293b;
        }
        #innerTable::item { padding: 5px 10px; border: none; }
        #innerTable::item:selected { background: #1e3a5f; }

        QScrollBar:vertical {
            background: #0b1120; width: 6px; border-radius: 3px;
        }
        QScrollBar::handle:vertical { background: #1e3a5f; border-radius: 3px; }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """

    # ── HELPERS ──────────────────────────────────────
    def _cargar_productos(self):
        self.cbo_producto.clear()
        self._prod_unidad.clear()
        for p in listar_productos("", incluir_inactivos=False):
            unidad = (getattr(p, "unidad", "") or "").strip() or "und"
            self._prod_unidad[p.id] = unidad
            self.cbo_producto.addItem(f"{p.codigo} - {p.nombre} ({unidad})", p.id)

    def _row_color(self, tipo: str) -> tuple[str, str]:
        t = (tipo or "").upper().strip()
        if t == "ENTRADA":
            return "#4ade80", "#0d2b1a"
        if t == "VENTA":
            return "#f87171", "#2b0d0d"
        if t == "ANULACION":
            return "#fbbf24", "#2b1e05"
        return "#e2e8f0", "#0b1120"

    def _set_resumen(self, unidad: str, rows: list) -> None:
        tot_entradas = tot_ventas = tot_anul = 0.0
        for r in rows:
            t = (r.tipo or "").upper().strip()
            c = float(r.cantidad or 0.0)
            if t == "ENTRADA":
                tot_entradas += c
            elif t == "VENTA":
                tot_ventas += abs(c)
            elif t == "ANULACION":
                tot_anul += c

        neto = tot_entradas - tot_ventas + tot_anul
        saldo_final = float(rows[-1].saldo) if rows else 0.0

        self._set_card(self._card_entradas, f"{_fmt_qty(tot_entradas)} {unidad}")
        self._set_card(self._card_ventas, f"{_fmt_qty(tot_ventas)} {unidad}")
        self._set_card(self._card_anul, f"{_fmt_qty(tot_anul)} {unidad}")
        self._set_card(self._card_neto, f"{_fmt_qty(neto)} {unidad}")
        self._set_card(self._card_saldo, f"{_fmt_qty(saldo_final)} {unidad}")

    # ── CARGA ─────────────────────────────────────────
    def keyPressEvent(self, event):
        from PySide6.QtCore import Qt

        key = event.key()
        mod = event.modifiers()
        if key == Qt.Key_F5:
            self.cargar()
        elif key == Qt.Key_Escape:
            self.close()
        else:
            super().keyPressEvent(event)

    def cargar(self):
        if self.cbo_producto.count() == 0:
            self.table.setRowCount(0)
            self.lbl_info.setText("Saldo inicial: 0")
            self._set_resumen("und", [])
            return

        product_id = int(self.cbo_producto.currentData())
        unidad = self._prod_unidad.get(product_id, "und")

        self.table.setHorizontalHeaderLabels(
            [
                "Fecha",
                "Tipo",
                "Referencia",
                f"Cantidad ({unidad})",
                "Precio",
                "Subtotal",
                f"Saldo ({unidad})",
            ]
        )

        d1 = self.dt_desde.date().toPython()
        d2 = self.dt_hasta.date().toPython()
        desde = datetime.combine(d1, time(0, 0, 0))
        hasta = datetime.combine(d2, time(23, 59, 59))

        data = obtener_kardex(product_id=product_id, desde=desde, hasta=hasta)
        saldo_inicial = float(data["saldo_inicial"] or 0.0)
        rows = data["rows"]
        advertencias = data.get("advertencias", [])

        info_txt = (
            f"Saldo inicial (antes del rango): {_fmt_qty(saldo_inicial)} {unidad}"
        )
        if advertencias:
            info_txt += "\n" + "\n".join(advertencias)
        self.lbl_info.setText(info_txt)

        self.table.blockSignals(True)
        self.table.setRowCount(len(rows))

        font_tipo = QFont()
        font_tipo.setBold(True)

        for i, r in enumerate(rows):
            self.table.setRowHeight(i, 32)
            tipo_txt = (r.tipo or "").upper().strip()
            color_txt, color_bg = self._row_color(tipo_txt)

            it_fecha = _item(fmt_fecha(r.fecha))
            it_tipo = _item(tipo_txt)
            it_ref = _item(r.referencia or "")
            it_cant = _item(_fmt_qty(r.cantidad), Qt.AlignRight | Qt.AlignVCenter)
            it_precio = _item(_fmt_cop(r.precio), Qt.AlignRight | Qt.AlignVCenter)
            it_sub = _item(_fmt_cop(r.subtotal), Qt.AlignRight | Qt.AlignVCenter)
            it_saldo = _item(_fmt_qty(r.saldo), Qt.AlignRight | Qt.AlignVCenter)

            it_tipo.setFont(font_tipo)
            it_tipo.setForeground(QBrush(QColor(color_txt)))
            it_ref.setToolTip(r.referencia or "")

            bg = QBrush(QColor(color_bg))
            for it in [it_fecha, it_tipo, it_ref, it_cant, it_precio, it_sub, it_saldo]:
                it.setBackground(bg)

            self.table.setItem(i, 0, it_fecha)
            self.table.setItem(i, 1, it_tipo)
            self.table.setItem(i, 2, it_ref)
            self.table.setItem(i, 3, it_cant)
            self.table.setItem(i, 4, it_precio)
            self.table.setItem(i, 5, it_sub)
            self.table.setItem(i, 6, it_saldo)

        self.table.blockSignals(False)
        self.table.resizeRowsToContents()
        self._set_resumen(unidad, rows)

    # ── EXPORTAR ─────────────────────────────────────
    def _kardex_para_exportar(self):
        if self.cbo_producto.count() == 0:
            return None
        product_id = int(self.cbo_producto.currentData())
        unidad = self._prod_unidad.get(product_id, "und")
        d1 = self.dt_desde.date().toPython()
        d2 = self.dt_hasta.date().toPython()
        desde = datetime.combine(d1, time(0, 0, 0))
        hasta = datetime.combine(d2, time(23, 59, 59))
        data = obtener_kardex(product_id=product_id, desde=desde, hasta=hasta)
        saldo_inicial = float(data["saldo_inicial"] or 0.0)
        rows = data["rows"]
        tot_entradas = tot_ventas = tot_anul = 0.0
        for r in rows:
            t = (r.tipo or "").upper().strip()
            c = float(r.cantidad or 0.0)
            if t == "ENTRADA":
                tot_entradas += c
            elif t == "VENTA":
                tot_ventas += abs(c)
            elif t == "ANULACION":
                tot_anul += c
        neto = tot_entradas - tot_ventas + tot_anul
        saldo_final = float(rows[-1].saldo) if rows else 0.0
        return {
            "product_id": product_id,
            "producto_text": self.cbo_producto.currentText(),
            "unidad": unidad,
            "desde": d1,
            "hasta": d2,
            "saldo_inicial": saldo_inicial,
            "rows": rows,
            "advertencias": data.get("advertencias", []),
            "tot_entradas": tot_entradas,
            "tot_ventas": tot_ventas,
            "tot_anul": tot_anul,
            "neto": neto,
            "saldo_final": saldo_final,
        }

    def exportar_pdf(self):
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
        except ImportError:
            QMessageBox.critical(
                self,
                "Dependencia faltante",
                "Instala reportlab:\n\npip install reportlab",
            )
            return
        try:
            payload = self._kardex_para_exportar()
            if not payload:
                QMessageBox.information(self, "Kardex", "No hay datos para exportar.")
                return
            producto = payload["producto_text"]
            d1, d2 = payload["desde"], payload["hasta"]
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar PDF", f"kardex_{d1}_a_{d2}.pdf", "PDF (*.pdf)"
            )
            if not path:
                return
            unidad = payload["unidad"]
            saldo_inicial = payload["saldo_inicial"]
            rows = payload["rows"]
            advertencias = payload["advertencias"]
            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter

            def header_page():
                nonlocal y
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2 * cm, y, "Kardex por Producto")
                y -= 0.7 * cm
                c.setFont("Helvetica", 10)
                c.drawString(2 * cm, y, f"Producto: {producto}")
                y -= 0.5 * cm
                c.drawString(2 * cm, y, f"Rango: {d1} a {d2}")
                y -= 0.5 * cm
                c.drawString(
                    2 * cm, y, f"Saldo inicial: {_fmt_qty(saldo_inicial)} {unidad}"
                )
                y -= 0.5 * cm
                if advertencias:
                    c.setFont("Helvetica-Oblique", 8)
                    for adv in advertencias:
                        c.drawString(2 * cm, y, adv[:110])
                        y -= 0.4 * cm
                    c.setFont("Helvetica", 10)
                    y -= 0.2 * cm
                c.setFont("Helvetica-Bold", 9)
                c.drawString(2 * cm, y, "Fecha")
                c.drawString(5.2 * cm, y, "Tipo")
                c.drawString(7.3 * cm, y, "Referencia")
                c.drawRightString(14.2 * cm, y, f"Cant ({unidad})")
                c.drawRightString(16.7 * cm, y, "Precio")
                c.drawRightString(19.5 * cm, y, "Subtotal")
                y -= 0.35 * cm
                c.line(2 * cm, y, 19.5 * cm, y)
                y -= 0.35 * cm
                c.setFont("Helvetica", 9)

            def wrap_text(text, max_width):
                words = (text or "").split()
                if not words:
                    return [""]
                lines, line = [], words[0]
                for w2 in words[1:]:
                    test = f"{line} {w2}"
                    if c.stringWidth(test, "Helvetica", 9) <= max_width:
                        line = test
                    else:
                        lines.append(line)
                        line = w2
                lines.append(line)
                return lines

            y = h - 2 * cm
            header_page()
            ref_col_width = (14.2 - 7.3) * cm - 0.3 * cm
            for r in rows:
                ref_lines = wrap_text(r.referencia or "", ref_col_width)
                needed = max(1, len(ref_lines)) * 0.42 * cm + 0.05 * cm
                if y < 2 * cm + needed:
                    c.showPage()
                    y = h - 2 * cm
                    header_page()
                fecha_txt = str(fmt_fecha(r.fecha) or "")[:16]
                tipo_txt = (r.tipo or "").upper().strip()
                c.drawString(2 * cm, y, fecha_txt)
                c.drawString(5.2 * cm, y, tipo_txt[:10])
                yy = y
                for line in ref_lines:
                    c.drawString(7.3 * cm, yy, line)
                    yy -= 0.42 * cm
                c.drawRightString(14.2 * cm, y, _fmt_qty(r.cantidad))
                c.drawRightString(16.7 * cm, y, _fmt_cop(r.precio))
                c.drawRightString(19.5 * cm, y, _fmt_cop(r.subtotal))
                y -= needed
            c.save()
            QMessageBox.information(self, "OK", "PDF exportado correctamente.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self,
                "Dependencia faltante",
                "Instala openpyxl:\n\npip install openpyxl",
            )
            return
        try:
            payload = self._kardex_para_exportar()
            if not payload:
                QMessageBox.information(self, "Kardex", "No hay datos para exportar.")
                return
            d1, d2 = payload["desde"], payload["hasta"]
            path, _ = QFileDialog.getSaveFileName(
                self, "Guardar Excel", f"kardex_{d1}_a_{d2}.xlsx", "Excel (*.xlsx)"
            )
            if not path:
                return
            unidad, rows = payload["unidad"], payload["rows"]
            advertencias = payload["advertencias"]
            wb = Workbook()
            ws = wb.active
            ws.title = "Kardex"
            ws["A1"] = "Kardex por Producto"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = "Producto:"
            ws["B2"] = payload["producto_text"]
            ws["A3"] = "Rango:"
            ws["B3"] = f"{d1} a {d2}"
            ws["A4"] = "Saldo inicial:"
            ws["B4"] = f"{_fmt_qty(payload['saldo_inicial'])} {unidad}"
            row_ptr = 6
            if advertencias:
                ws[f"A{row_ptr}"] = "⚠️ Advertencias:"
                ws[f"A{row_ptr}"].font = Font(bold=True, color="FF8800")
                row_ptr += 1
                for adv in advertencias:
                    ws[f"A{row_ptr}"] = adv
                    ws[f"A{row_ptr}"].font = Font(italic=True, color="FF8800")
                    row_ptr += 1
                row_ptr += 1
            ws[f"A{row_ptr}"] = "Resumen"
            ws[f"A{row_ptr}"].font = Font(bold=True)
            row_ptr += 1
            for label, val in [
                ("Entradas", payload["tot_entradas"]),
                ("Ventas", payload["tot_ventas"]),
                ("Anulaciones", payload["tot_anul"]),
                ("Neto", payload["neto"]),
                ("Saldo final", payload["saldo_final"]),
            ]:
                ws[f"A{row_ptr}"] = label
                ws[f"B{row_ptr}"] = f"{_fmt_qty(val)} {unidad}"
                row_ptr += 1
            row_ptr += 1
            headers = [
                "Fecha",
                "Tipo",
                "Referencia",
                f"Cantidad ({unidad})",
                "Precio",
                "Subtotal",
                f"Saldo ({unidad})",
            ]
            for col, htxt in enumerate(headers, 1):
                ws.cell(row=row_ptr, column=col, value=htxt).font = Font(bold=True)
            align_right = Alignment(horizontal="right")
            for i, r in enumerate(rows, start=row_ptr + 1):
                ws.cell(i, 1, fmt_fecha(r.fecha))
                ws.cell(i, 2, (r.tipo or "").upper().strip())
                ws.cell(i, 3, r.referencia or "")
                ws.cell(i, 4, _fmt_qty(r.cantidad)).alignment = align_right
                ws.cell(i, 5, _fmt_cop(r.precio)).alignment = align_right
                ws.cell(i, 6, _fmt_cop(r.subtotal)).alignment = align_right
                ws.cell(i, 7, _fmt_qty(r.saldo)).alignment = align_right
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 18
            ws.column_dimensions["C"].width = 55
            wb.save(path)
            QMessageBox.information(self, "OK", "Excel exportado correctamente.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
